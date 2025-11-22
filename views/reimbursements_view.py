"""
Reimbursements Tab - Track expenses awaiting reimbursement

Purpose:
    - Track work travel expenses, friend loans, and other temporary out-of-pocket costs
    - Filter by trip/event tags (e.g., "Whispers25", "NYC24")
    - Export filtered data for expense reports
    - Separate from main budget calculations (money is expected to be reimbursed)

Layout:
    - Top row: Tag filter (left), action buttons (middle), future charts area (right)
    - Bottom: Full-width sortable/editable table with all reimbursements
"""

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QListWidget, QListWidgetItem, QPushButton,
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QAbstractItemView, QFrame, QSpacerItem, QSizePolicy, QMessageBox, QFileDialog)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor
from themes import theme_manager
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font as ExcelFont, Alignment
from widgets.chart_widget import (ReimbursementStatsWidget, ReimbursementProgressWidget,
                                  ReimbursementDotPlotWidget, ReimbursementHeatmapWidget)


class ReimbursementsView(QWidget):
    """Reimbursements tab for tracking expenses awaiting reimbursement"""

    def __init__(self, transaction_manager, parent=None):
        super().__init__(parent)
        self.transaction_manager = transaction_manager
        self.current_filter_tag = "All"  # Track which tag is selected

        # Change tracking for Save/Delete functionality
        self.pending_changes = {}  # {row: {column: new_value}}
        self.rows_to_delete = set()  # Set of row indices marked for deletion
        self.original_values = {}  # {row: {column: original_value}}
        self.reimbursement_ids = {}  # {row: reimbursement_id}
        self.loading_data = False  # Flag to prevent cellChanged signal during table load
        self.last_dot_plot_state = None  # Track if dot plot is showing to prevent flashing

        self.init_ui()

        # Connect to theme changes
        theme_manager.theme_changed.connect(self.on_theme_changed)

    def showEvent(self, event):
        """Called when widget is shown - update layout after sizing is known

        This Qt event handler is triggered when the Reimbursements tab becomes visible.
        At this point, the widget has been laid out and has actual dimensions,
        so we can accurately calculate whether to show the dot plot widget.

        Called during: Tab switching, app startup (if Reimbursements is default tab)
        """
        super().showEvent(event)
        # Update charts layout now that we have actual dimensions
        if hasattr(self, 'charts_layout'):
            self.update_charts_layout()

    def resizeEvent(self, event):
        """Called when widget is resized - update layout dynamically

        This Qt event handler is triggered when the window size changes.
        We use this to show/hide the dot plot widget based on available space.

        Called during: Window resize, maximize/restore, screen resolution change
        """
        super().resizeEvent(event)
        # Update charts layout based on new size
        if hasattr(self, 'charts_layout'):
            self.update_charts_layout()

    def init_ui(self):
        """Initialize the UI layout"""
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)

        # TOP ROW: Tag filter, buttons, and future charts area
        top_row_layout = QHBoxLayout()
        top_row_layout.setSpacing(10)

        # Tag filter (left side of top row)
        tag_filter = self.create_tag_sidebar()
        top_row_layout.addWidget(tag_filter)

        # Action buttons (middle of top row)
        button_row = self.create_button_row()
        top_row_layout.addLayout(button_row)

        # Charts area (right side of top row)
        charts_layout = self.create_charts_area()
        top_row_layout.addLayout(charts_layout, stretch=1)

        main_layout.addLayout(top_row_layout)

        # BOTTOM: Full-width table (stretches to fill remaining space)
        self.table = self.create_table()
        main_layout.addWidget(self.table, stretch=1)

        self.setLayout(main_layout)

        # Load initial data
        self.refresh()

    def create_tag_sidebar(self):
        """Create the tag filter list (compact for top row)"""
        sidebar_frame = QFrame()
        sidebar_frame.setMinimumWidth(200)
        sidebar_frame.setMaximumWidth(250)
        # Height matches button column: 4 buttons * 35px + 3 gaps * 10px = 170px
        sidebar_frame.setFixedHeight(170)
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setSpacing(5)
        sidebar_layout.setContentsMargins(5, 5, 5, 5)

        colors = theme_manager.get_colors()

        # Title label
        title_label = QLabel("Filter by Tag:")
        title_label.setFont(theme_manager.get_font("subtitle"))
        title_label.setStyleSheet(f"color: {colors['primary']}; font-weight: bold;")
        sidebar_layout.addWidget(title_label)

        # Tag list widget
        self.tag_list = QListWidget()
        self.tag_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.tag_list.itemClicked.connect(self.on_tag_selected)

        sidebar_layout.addWidget(self.tag_list)

        sidebar_frame.setLayout(sidebar_layout)
        sidebar_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {colors['surface_variant']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
            }}
        """)

        return sidebar_frame

    def create_button_row(self):
        """Create horizontal row of action buttons"""
        button_layout = QVBoxLayout()
        button_layout.setSpacing(10)

        colors = theme_manager.get_colors()

        # Add button - opens Add Transaction dialog with Reimbursements pre-selected
        self.add_button = QPushButton("Add")
        self.add_button.setFixedHeight(35)
        self.add_button.setFixedWidth(100)
        self.add_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.add_button.clicked.connect(self.open_add_reimbursement_dialog)
        button_layout.addWidget(self.add_button)

        # Save button - commits all pending changes and deletions to database
        self.save_button = QPushButton("Save")
        self.save_button.setFixedHeight(35)
        self.save_button.setFixedWidth(100)
        self.save_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.save_button.clicked.connect(self.save_all_changes)
        button_layout.addWidget(self.save_button)

        # Delete button - marks selected rows for deletion (red text)
        self.delete_button = QPushButton("Delete")
        self.delete_button.setFixedHeight(35)
        self.delete_button.setFixedWidth(100)
        self.delete_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.delete_button.clicked.connect(self.mark_for_deletion)
        button_layout.addWidget(self.delete_button)

        # Export button - exports selected or all reimbursements to Excel
        self.export_button = QPushButton("Export")
        self.export_button.setFixedHeight(35)
        self.export_button.setFixedWidth(100)
        self.export_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.export_button.clicked.connect(self.export_reimbursements)
        button_layout.addWidget(self.export_button)

        # Add stretch to push buttons to top of their vertical column
        button_layout.addStretch()

        # Apply button styling
        self.apply_button_theme()

        return button_layout

    def create_charts_area(self):
        """Create the charts visualization area with dynamic columns"""
        # Create a container widget that we can update later
        self.charts_container = QWidget()
        self.charts_layout = QHBoxLayout()
        self.charts_layout.setSpacing(10)
        self.charts_layout.setContentsMargins(0, 0, 0, 0)

        # Height should match button column: 4 buttons * 35px + 3 gaps * 10px = 170px
        self.CHART_HEIGHT = 170

        # Create all widgets once
        # Column 1: Stats widget (Total $ + Pie chart) - Fixed width and height
        self.stats_widget = ReimbursementStatsWidget(title="", parent=self)
        self.stats_widget.setFixedWidth(200)
        self.stats_widget.setFixedHeight(self.CHART_HEIGHT)

        # Column 2 (optional): Dot plot widget - Square, same height
        self.dot_plot_widget = ReimbursementDotPlotWidget(title="", parent=self)
        self.dot_plot_widget.setFixedSize(self.CHART_HEIGHT, self.CHART_HEIGHT)  # Square
        self.dot_plot_widget.hide()  # Hidden by default

        # Column 3: Progress bars widget (Submitted & Reimbursed) - Fixed width and height
        self.progress_widget = ReimbursementProgressWidget(title="", parent=self)
        self.progress_widget.setFixedWidth(180)
        self.progress_widget.setFixedHeight(self.CHART_HEIGHT)

        # Column 4: Heatmap widget (Tag × Category) - Fixed size, snug around content
        self.heatmap_widget = ReimbursementHeatmapWidget(title="", parent=self)
        self.heatmap_widget.setFixedWidth(250)  # Snug fit for 2×2 grid
        self.heatmap_widget.setFixedHeight(self.CHART_HEIGHT)

        # Initial layout (without dot plot)
        self.update_charts_layout()

        self.charts_container.setLayout(self.charts_layout)

        # Return container as a layout wrapper
        wrapper_layout = QHBoxLayout()
        wrapper_layout.addWidget(self.charts_container, stretch=1)
        return wrapper_layout

    def update_charts_layout(self):
        """Update charts layout based on available space

        This method implements ADAPTIVE LAYOUT for the dot plot widget:
        - Calculates available horizontal space dynamically
        - Shows dot plot only when there's enough room
        - Uses state tracking to prevent unnecessary rebuilds (avoid flashing)

        Layout order (left to right):
        1. Stats widget (always shown, 200px)
        2. Dot plot widget (adaptive, 170px, only when space available)
        3. Progress widget (always shown, 180px)
        4. Heatmap widget (always shown, 250px)
        5. Stretch spacer (pushes widgets left)

        State tracking prevents flashing:
        - self.last_dot_plot_state tracks if dot plot was shown last time
        - Only rebuilds layout when show/hide state changes
        - Prevents 5+ rebuilds during tab initialization
        """
        # Get the actual width of the charts container
        container_width = self.charts_container.width() if hasattr(self, 'charts_container') else self.width()

        # Pull actual widths from widgets dynamically (not hardcoded)
        # This ensures layout adapts to different screen sizes
        stats_width = self.stats_widget.width()
        progress_width = self.progress_widget.width()
        dot_plot_width = self.dot_plot_widget.width()
        heatmap_width = self.heatmap_widget.width()
        spacing = self.charts_layout.spacing()

        # Calculate space needed if dot plot is included
        # Formula: stats + dot + progress + heatmap + (4 gaps between widgets)
        space_with_dot = stats_width + dot_plot_width + progress_width + heatmap_width + (4 * spacing)

        # Determine if we should show dot plot
        # Use 30px tolerance to account for layout calculation variations
        tolerance = 30
        show_dot_plot = container_width >= (space_with_dot - tolerance)

        # Only rebuild layout if dot plot visibility state changed
        # This prevents flashing during initialization (5+ calls to this method)
        if show_dot_plot == self.last_dot_plot_state:
            return  # No change needed, skip rebuild

        # Update state tracking
        self.last_dot_plot_state = show_dot_plot

        # Clear existing layout (remove all widgets)
        while self.charts_layout.count():
            item = self.charts_layout.takeAt(0)
            if item.widget():
                item.widget().setParent(None)

        # Rebuild layout in correct order
        # Stats widget (always first)
        self.charts_layout.addWidget(self.stats_widget)

        # Dot plot widget (adaptive - only add if we have space)
        if show_dot_plot:
            self.dot_plot_widget.show()
            self.charts_layout.addWidget(self.dot_plot_widget)
        else:
            self.dot_plot_widget.hide()

        # Progress widget (always shown)
        self.charts_layout.addWidget(self.progress_widget)

        # Heatmap widget (always shown, fixed width - no stretch factor)
        self.charts_layout.addWidget(self.heatmap_widget)

        # Add stretch spacer to push all widgets to the left
        # Without this, heatmap would expand to fill remaining space
        self.charts_layout.addStretch(1)

    def create_table(self):
        """Create the main reimbursements table"""
        table = QTableWidget()

        # Column headers
        headers = ["Date", "Status", "Amount", "Tag", "Category", "Notes"]
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        # Enable sorting
        table.setSortingEnabled(True)

        # Selection mode: single click selects row, double click edits cell
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)  # Allow multi-select with Ctrl
        table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)  # Double-click to edit

        # Resize modes
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Date
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Status
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Amount
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Tag
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)  # Category
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)  # Notes (stretch to fill)

        # Make header clickable for sorting
        header.setSectionsClickable(True)

        # Connect cell change signal to track edits
        table.cellChanged.connect(self.on_cell_changed)

        # Apply table styling
        colors = theme_manager.get_colors()
        table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {colors['surface']};
                border: 1px solid {colors['border']};
                gridline-color: {colors['border']};
            }}
            QTableWidget::item {{
                padding: 5px;
            }}
            QTableWidget::item:selected {{
                background-color: {colors['accent']};
            }}
            QHeaderView::section {{
                background-color: {colors['surface_variant']};
                color: {colors['text_primary']};
                padding: 5px;
                border: 1px solid {colors['border']};
                font-weight: bold;
            }}
            QHeaderView::section:hover {{
                background-color: {colors['primary']};
            }}
        """)

        return table

    def apply_button_theme(self):
        """Apply theme styling to buttons"""
        colors = theme_manager.get_colors()

        button_style = f"""
            QPushButton {{
                background-color: {colors['primary']};
                color: {colors['background']};
                border: none;
                border-radius: 4px;
                padding: 5px 10px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {colors['accent']};
            }}
            QPushButton:pressed {{
                background-color: {colors['surface_variant']};
            }}
        """

        self.add_button.setStyleSheet(button_style)
        self.save_button.setStyleSheet(button_style)
        self.delete_button.setStyleSheet(button_style)
        self.export_button.setStyleSheet(button_style)

    def apply_table_theme(self):
        """Apply theme styling to table"""
        colors = theme_manager.get_colors()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {colors['surface']};
                border: 1px solid {colors['border']};
                gridline-color: {colors['border']};
            }}
            QTableWidget::item {{
                color: {colors['text_primary']};
                padding: 5px;
            }}
            QTableWidget::item:selected {{
                background-color: {colors['primary']};
                color: {colors['background']};
            }}
            QHeaderView::section {{
                background-color: {colors['surface_variant']};
                color: {colors['text_primary']};
                padding: 5px;
                border: 1px solid {colors['border']};
                font-weight: bold;
            }}
        """)

    def apply_sidebar_theme(self):
        """Apply theme styling to tag sidebar"""
        colors = theme_manager.get_colors()

        # Update sidebar frame
        if hasattr(self, 'tag_list'):
            parent_frame = self.tag_list.parent()
            if parent_frame:
                parent_frame.setStyleSheet(f"""
                    QFrame {{
                        background-color: {colors['surface_variant']};
                        border: 1px solid {colors['border']};
                        border-radius: 4px;
                    }}
                """)

            # Update tag list styling
            self.tag_list.setStyleSheet(f"""
                QListWidget {{
                    background-color: {colors['background']};
                    border: 1px solid {colors['border']};
                    border-radius: 4px;
                }}
                QListWidget::item {{
                    color: {colors['text_primary']};
                    padding: 5px;
                }}
                QListWidget::item:hover {{
                    background-color: {colors['surface_variant']};
                }}
                QListWidget::item:selected {{
                    background-color: {colors['primary']};
                    color: {colors['background']};
                }}
            """)

        # Update title label if it exists
        for child in self.findChildren(QLabel):
            if child.text() == "Filter by Tag:":
                child.setStyleSheet(f"color: {colors['primary']}; font-weight: bold;")
                break

    def populate_tag_list(self):
        """Populate the tag filter list from database"""
        self.tag_list.clear()

        # Add "All" option
        all_item = QListWidgetItem("All")
        all_item.setData(Qt.ItemDataRole.UserRole, "All")
        self.tag_list.addItem(all_item)

        # Add "Other" option (for NULL/empty tags)
        other_item = QListWidgetItem("Other")
        other_item.setData(Qt.ItemDataRole.UserRole, "Other")
        self.tag_list.addItem(other_item)

        # Load unique tags from database
        try:
            from services.reimbursement_manager import ReimbursementManager
            rm = ReimbursementManager()
            try:
                tags = rm.get_unique_locations()
                for tag in sorted(tags):
                    if tag:  # Only add non-empty tags
                        tag_item = QListWidgetItem(tag)
                        tag_item.setData(Qt.ItemDataRole.UserRole, tag)
                        self.tag_list.addItem(tag_item)
            finally:
                rm.close()
        except Exception as e:
            print(f"Error loading tags: {e}")

        # Select "All" by default
        self.tag_list.setCurrentRow(0)

    def populate_table(self):
        """Populate the table with reimbursements (filtered by selected tag)"""
        # Set loading flag to block cellChanged signal
        self.loading_data = True

        # Clear tracking data
        self.pending_changes = {}
        self.rows_to_delete = set()
        self.original_values = {}
        self.reimbursement_ids = {}

        # Clear existing rows
        self.table.setRowCount(0)

        try:
            from services.reimbursement_manager import ReimbursementManager
            rm = ReimbursementManager()
            try:
                # Get all reimbursements (we'll filter in Python for now)
                all_reimbursements = rm.get_all_reimbursements(sort_by="date", ascending=False)

                # Filter by selected tag for table display
                if self.current_filter_tag == "All":
                    reimbursements = all_reimbursements
                elif self.current_filter_tag == "Other":
                    reimbursements = [r for r in all_reimbursements if not r.location or r.location.strip() == ""]
                else:
                    reimbursements = [r for r in all_reimbursements if r.location == self.current_filter_tag]

                # Update charts BEFORE populating table
                # Stats, Progress, and Dot Plot widgets use filtered data (respects tag selection)
                self.stats_widget.update_data(reimbursements)
                self.progress_widget.update_data(reimbursements)
                self.dot_plot_widget.update_data(reimbursements)
                # Heatmap widget always uses ALL data (ignores tag filter)
                self.heatmap_widget.update_data(all_reimbursements)

                # Populate table rows
                self.table.setRowCount(len(reimbursements))

                for row, reimbursement in enumerate(reimbursements):
                    # Store reimbursement ID for this row
                    self.reimbursement_ids[row] = reimbursement.id

                    # Initialize original values for this row
                    if row not in self.original_values:
                        self.original_values[row] = {}

                    # Date (editable)
                    date_str = reimbursement.date.strftime('%m/%d/%Y')
                    date_item = QTableWidgetItem(date_str)
                    self.original_values[row][0] = date_str
                    self.table.setItem(row, 0, date_item)

                    # Status (editable)
                    status_str = reimbursement.status_display
                    status_item = QTableWidgetItem(status_str)
                    self.original_values[row][1] = status_str
                    self.table.setItem(row, 1, status_item)

                    # Amount (editable) - store without $ sign for easier editing
                    amount_str = f"{reimbursement.amount:.2f}"
                    amount_item = QTableWidgetItem(amount_str)
                    self.original_values[row][2] = amount_str
                    self.table.setItem(row, 2, amount_item)

                    # Tag (editable)
                    tag_str = reimbursement.location or ""
                    tag_item = QTableWidgetItem(tag_str)
                    self.original_values[row][3] = tag_str
                    self.table.setItem(row, 3, tag_item)

                    # Category (editable)
                    category_str = reimbursement.category or ""
                    category_item = QTableWidgetItem(category_str)
                    self.original_values[row][4] = category_str
                    self.table.setItem(row, 4, category_item)

                    # Notes (editable)
                    notes_str = reimbursement.notes or ""
                    notes_item = QTableWidgetItem(notes_str)
                    self.original_values[row][5] = notes_str
                    self.table.setItem(row, 5, notes_item)

            finally:
                rm.close()

        except Exception as e:
            print(f"Error populating table: {e}")
        finally:
            # Clear loading flag to re-enable cellChanged signal
            self.loading_data = False

    def on_tag_selected(self, item):
        """Handle tag selection from sidebar"""
        self.current_filter_tag = item.data(Qt.ItemDataRole.UserRole)
        self.populate_table()

    def open_add_reimbursement_dialog(self):
        """Open Add Transaction dialog with Reimbursements mode pre-selected"""
        from views.dialogs.add_transaction_dialog import AddTransactionDialog

        dialog = AddTransactionDialog(self.transaction_manager, parent=self)

        # Pre-select "Reimbursements" in the transaction type dropdown
        reimbursements_index = dialog.mode_combo.findText("Reimbursements")
        if reimbursements_index >= 0:
            dialog.mode_combo.setCurrentIndex(reimbursements_index)

        # Show dialog and refresh table if accepted
        if dialog.exec():
            self.refresh()

    def on_cell_changed(self, row, column):
        """Track cell changes with visual highlighting (don't save immediately)"""
        try:
            # Ignore signals during data loading
            if self.loading_data:
                return

            # Get reimbursement ID for this row
            reimbursement_id = self.reimbursement_ids.get(row)
            if not reimbursement_id:
                return

            # Get the new value
            item = self.table.item(row, column)
            if not item:
                return

            new_value = item.text()

            # Check if value actually changed from original
            original_value = self.original_values.get(row, {}).get(column, "")
            if new_value == original_value:
                # Value reverted to original - remove from pending changes
                if row in self.pending_changes and column in self.pending_changes[row]:
                    del self.pending_changes[row][column]
                    if not self.pending_changes[row]:  # Remove row if no changes left
                        del self.pending_changes[row]
                    # Clear highlight
                    item.setBackground(Qt.GlobalColor.transparent)
                return

            # Store the pending change
            if row not in self.pending_changes:
                self.pending_changes[row] = {}

            self.pending_changes[row][column] = new_value

            # Mark cell as changed visually with subtle theme-based color
            colors = theme_manager.get_colors()
            warning_color = colors.get('warning', '#FACC15')
            surface_color = colors.get('surface', '#161616')

            # Blend warning color with surface for subtle highlight (25% warning, 75% surface)
            warning_qcolor = QColor(warning_color)
            surface_qcolor = QColor(surface_color)

            blend_r = int(surface_qcolor.red() * 0.75 + warning_qcolor.red() * 0.25)
            blend_g = int(surface_qcolor.green() * 0.75 + warning_qcolor.green() * 0.25)
            blend_b = int(surface_qcolor.blue() * 0.75 + warning_qcolor.blue() * 0.25)
            blended_color = QColor(blend_r, blend_g, blend_b)

            item.setBackground(blended_color)

        except Exception as e:
            print(f"Error tracking cell change: {e}")

    def mark_for_deletion(self):
        """Mark selected rows for deletion (red text)"""
        selected_rows = set()
        for item in self.table.selectedItems():
            selected_rows.add(item.row())

        if not selected_rows:
            QMessageBox.information(self, "No Selection", "Please select rows to mark for deletion")
            return

        # Mark rows for deletion (no confirmation popup - confirmation happens at save)
        for row in selected_rows:
            self.rows_to_delete.add(row)
            # Visual indication - make entire row text red
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setForeground(QColor("#EF4444"))  # Red text color
                    item.setBackground(Qt.GlobalColor.transparent)  # Clear background

    def save_all_changes(self):
        """Apply all pending changes and deletions to the database"""
        if not self.pending_changes and not self.rows_to_delete:
            QMessageBox.information(self, "No Changes", "No changes to save")
            return

        # Build detailed change list for confirmation
        change_summary = []

        # List edits
        for row, changes in self.pending_changes.items():
            reimbursement_id = self.reimbursement_ids.get(row)
            if not reimbursement_id:
                continue

            for column, new_value in changes.items():
                field_names = {0: "Date", 1: "Status", 2: "Amount", 3: "Tag", 4: "Category", 5: "Notes"}
                field_name = field_names.get(column, f"Column {column}")
                original_value = self.original_values.get(row, {}).get(column, "")

                change_summary.append(
                    f"ID {reimbursement_id}: Change {field_name} from '{original_value}' to '{new_value}'"
                )

        # List deletions
        for row in self.rows_to_delete:
            reimbursement_id = self.reimbursement_ids.get(row)
            if reimbursement_id:
                change_summary.append(f"ID {reimbursement_id}: DELETE")

        # Show confirmation dialog
        if change_summary:
            message = "Proposed Changes:\n\n" + "\n".join(change_summary)
        else:
            message = "No changes to apply"

        reply = QMessageBox.question(self, "Confirm Changes", message,
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            from services.reimbursement_manager import ReimbursementManager
            rm = ReimbursementManager()

            try:
                # Map status display names back to state enum values
                status_to_state_map = {
                    "Pending Submission": "pending",
                    "Awaiting Payment": "submitted",
                    "Reimbursed": "reimbursed",
                    "Partially Reimbursed": "partial",
                    "Denied": "denied"
                }

                # Apply edits first
                updated_count = 0
                for row, changes in self.pending_changes.items():
                    reimbursement_id = self.reimbursement_ids.get(row)
                    if not reimbursement_id:
                        continue

                    # Build update data
                    update_data = {}

                    for column, new_value in changes.items():
                        if column == 0:  # Date
                            try:
                                new_date = datetime.strptime(new_value, '%m/%d/%Y').date()
                                update_data['date'] = new_date
                            except Exception as e:
                                QMessageBox.warning(self, "Invalid Date", f"Invalid date format in row {row+1}: {new_value}\nUse MM/DD/YYYY format")
                                return
                        elif column == 1:  # Status
                            new_state = status_to_state_map.get(new_value)
                            if new_state:
                                update_data['state'] = new_state
                            else:
                                QMessageBox.warning(self, "Invalid Status", f"Invalid status in row {row+1}: {new_value}")
                                return
                        elif column == 2:  # Amount
                            try:
                                new_amount = float(new_value.replace('$', '').replace(',', ''))
                                update_data['amount'] = abs(new_amount)  # Always positive
                            except Exception as e:
                                QMessageBox.warning(self, "Invalid Amount", f"Invalid amount in row {row+1}: {new_value}")
                                return
                        elif column == 3:  # Tag
                            update_data['location'] = new_value if new_value.strip() else None
                        elif column == 4:  # Category
                            if not new_value.strip():
                                QMessageBox.warning(self, "Category Required", f"Category is required (row {row+1})")
                                return
                            update_data['category'] = new_value.strip()
                        elif column == 5:  # Notes
                            update_data['notes'] = new_value

                    # Update reimbursement
                    if update_data:
                        result = rm.update_reimbursement(reimbursement_id, update_data)
                        if result:
                            updated_count += 1
                        else:
                            QMessageBox.warning(self, "Update Failed", f"Failed to update reimbursement ID {reimbursement_id}")
                            return

                # Apply deletions
                deleted_count = 0
                for row in self.rows_to_delete:
                    reimbursement_id = self.reimbursement_ids.get(row)
                    if reimbursement_id:
                        if rm.delete_reimbursement(reimbursement_id):
                            deleted_count += 1

                # Show success message
                success_msg = []
                if updated_count > 0:
                    success_msg.append(f"Updated {updated_count} reimbursement(s)")
                if deleted_count > 0:
                    success_msg.append(f"Deleted {deleted_count} reimbursement(s)")

                if success_msg:
                    QMessageBox.information(self, "Success", "\n".join(success_msg))

                # Refresh table
                self.refresh()

            finally:
                rm.close()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving changes: {str(e)}")
            import traceback
            traceback.print_exc()

    def export_reimbursements(self):
        """Export selected (or all) reimbursements to Excel file"""
        try:
            # Get selected rows
            selected_rows = set()
            for item in self.table.selectedItems():
                selected_rows.add(item.row())

            # If no selection, use all rows in table
            if not selected_rows:
                row_count = self.table.rowCount()
                if row_count == 0:
                    QMessageBox.warning(self, "No Data", "You must select at least 1 reimbursement to export")
                    return
                selected_rows = set(range(row_count))

            # Get reimbursement data for selected rows
            export_data = []
            for row in sorted(selected_rows):
                reimbursement_id = self.reimbursement_ids.get(row)
                if not reimbursement_id:
                    continue

                # Get data from table cells
                date_item = self.table.item(row, 0)
                status_item = self.table.item(row, 1)
                amount_item = self.table.item(row, 2)
                tag_item = self.table.item(row, 3)
                category_item = self.table.item(row, 4)
                notes_item = self.table.item(row, 5)

                export_data.append({
                    'amount': amount_item.text() if amount_item else "",
                    'tag': tag_item.text() if tag_item else "",
                    'category': category_item.text() if category_item else "",
                    'notes': notes_item.text() if notes_item else "",
                    'status': status_item.text() if status_item else ""
                })

            if not export_data:
                QMessageBox.warning(self, "No Data", "No reimbursements to export")
                return

            # Generate default filename
            today = datetime.now()
            date_str = today.strftime('%m%d%y')  # Format: MMDDYY (e.g., 111925)

            # Include tag in filename if not "All" or empty
            if self.current_filter_tag and self.current_filter_tag not in ["All", "Other", ""]:
                default_filename = f"Reimbursements_{self.current_filter_tag}_{date_str}.xlsx"
            else:
                default_filename = f"Reimbursements_{date_str}.xlsx"

            # Get default directory (same as database location)
            default_dir = Path(__file__).parent.parent  # Go up to BudgetApp directory
            default_path = default_dir / default_filename

            # Open file save dialog
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Reimbursements",
                str(default_path),
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if not file_path:
                return  # User cancelled

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reimbursements"

            # Write headers in row 1 (A1, B1, C1, D1, E1)
            headers = ["Amount", "Tag", "Category", "Notes", "Status"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = ExcelFont(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='top')

            # Write data starting from row 2
            for row_idx, data in enumerate(export_data, start=2):
                ws.cell(row=row_idx, column=1, value=data['amount'])
                ws.cell(row=row_idx, column=2, value=data['tag'])
                ws.cell(row=row_idx, column=3, value=data['category'])
                ws.cell(row=row_idx, column=4, value=data['notes'])
                ws.cell(row=row_idx, column=5, value=data['status'])

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(file_path)

            # Show success message
            QMessageBox.information(
                self,
                "Export Successful",
                f"Exported {len(export_data)} reimbursement(s) to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting reimbursements:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def refresh(self):
        """Refresh the entire view (tag list and table)"""
        self.populate_tag_list()
        self.populate_table()

    def on_theme_changed(self):
        """Handle theme changes"""
        self.apply_button_theme()
        self.apply_table_theme()
        self.apply_sidebar_theme()
        # Refresh charts with new theme colors
        if hasattr(self, 'charts_layout'):
            self.update_charts_layout()
